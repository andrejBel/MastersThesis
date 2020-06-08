from pathlib import Path
from typing import List

from openpyxl import Workbook, load_workbook


class ExcelLogger:
    path_to_workbook = None
    work_sheet = None
    wb = None
    buffer = None
    log = False

    @staticmethod
    def init_logger(path_to_workbook):
        ExcelLogger.path_to_workbook = path_to_workbook
        if ExcelLogger.path_to_workbook is not None:
            if Path(ExcelLogger.path_to_workbook).exists():
                ExcelLogger.wb = load_workbook(path_to_workbook)
            else:
                ExcelLogger.wb = Workbook()
            ExcelLogger.work_sheet = ExcelLogger.wb.active
            ExcelLogger.buffer = []
            ExcelLogger.log = True
        else:
            ExcelLogger.wb = None
            ExcelLogger.work_sheet = None
            ExcelLogger.buffer = None
            ExcelLogger.log = False

    @staticmethod
    def set_logging(value):
        ExcelLogger.log = value

    @staticmethod
    def write_into_sheet(iterable: List, force_log=False):
        if ExcelLogger.work_sheet is not None:
            if force_log or ExcelLogger.log:
                ExcelLogger.work_sheet.append(iterable)
                ExcelLogger.wb.save(ExcelLogger.path_to_workbook)

    @staticmethod
    def extend_buffer(iterable, force_log=False):
        if ExcelLogger.buffer is not None:
            if force_log or ExcelLogger.log:
                ExcelLogger.buffer.extend(iterable)

    @staticmethod
    def append_buffer(iterable, force_log=False):
        if ExcelLogger.buffer is not None:
            if force_log or ExcelLogger.log:
                ExcelLogger.buffer.append(iterable)

    @staticmethod
    def flush():
        if ExcelLogger.buffer is not None:
            ExcelLogger.write_into_sheet(ExcelLogger.buffer, True)
            ExcelLogger.buffer.clear()

    @staticmethod
    def append_to_workbook(to_append, path_to_workbook):
        if path_to_workbook is not None:
            wb: Workbook = None
            if Path(path_to_workbook).exists():
                wb = load_workbook(path_to_workbook)
            else:
                wb = Workbook()
            wb.active.append(to_append)
            wb.save(path_to_workbook)
