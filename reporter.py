from pathlib import Path
from typing import List

import numpy as np
from openpyxl import Workbook
from openpyxl import load_workbook

import constants
import datasets
import experiments
import models
from experiments import BasicTrainParametersClassifier, BasicTrainParametersAutoencoder, BasicTrainParametersTwoModels, \
    BasicTrainParametersAutoClassifier, ExperimentBase
from experiments import ExperimentResult, ExperimentClassifier, ExperimentAutoencoder, ExperimentAutoClassifier, \
    ExperimentAutoencoderAndClassifier


def clear_float(input):
    return str(input).replace('.', ',')


class Reporter():

    @staticmethod
    def create_classifier_report(experiment_results: List[ExperimentResult], path_to_workbook, experiment_type):
        wb: Workbook = None
        if Path(path_to_workbook).exists():
            wb = load_workbook(path_to_workbook)
        else:
            wb = Workbook()
        ws = wb.active
        ws.append(["Typ experimentu", experiment_type])
        ws.append(
            ['Vrstvy enkódera trénovateľné',
             "Percento trénovacích dát",
             "Počet epoch",
             "Najlepšia epocha",
             "Presnosť na trénovacom datasete",
             "Presnosť na testovacom datasete"])
        for result in experiment_results:
            train_history: BasicTrainParametersClassifier = result.train_history.train_parameters
            rate = train_history.train_data_rate
            encoder_trainable = train_history.autoencoder_layers_trainable_during_classifier_training
            encoder_trainable_str = "Áno" if encoder_trainable else "Nie"
            for model in result.train_history.train_history_model_list:
                if model.model_name == constants.Models.CLASSIFIER:
                    accuracy_on_train = -1
                    if constants.Metrics.ACCURACY in model.metrics:
                        accuracy_list = model.values[constants.Metrics.ACCURACY]
                        accuracy_on_train = accuracy_list[model.best_epoch - 1]
                    ws.append([encoder_trainable_str,
                               str(round(rate * 100, 0)).replace('.', ','),
                               model.stopped_epoch,
                               model.best_epoch,
                               str(accuracy_on_train).replace('.', ','),
                               str(model.monitor_best_value).replace('.', ',')])
        wb.save(path_to_workbook)

    @staticmethod
    def create_autoencoder_report(experiment_results: List[ExperimentResult], path_to_workbook, experiment_type):
        wb: Workbook = None
        if Path(path_to_workbook).exists():
            wb = load_workbook(path_to_workbook)
        else:
            wb = Workbook()
        ws = wb.active
        ws.append(["Typ experimentu", experiment_type])
        ws.append(
            ["Percento trénovacích dát",
             "Počet epoch",
             "Najlepšia epocha",
             "Rekonštrukčná chyba trénovací dataset",
             "Rekonštrukčná chyba testovací dataset"])
        for result in experiment_results:
            train_history: BasicTrainParametersAutoencoder = result.train_history.train_parameters
            rate = train_history.train_data_rate
            for model in result.train_history.train_history_model_list:
                if model.model_name == constants.Models.AUTOENCODER:
                    reconstruction_error_on_train = -1
                    if constants.Metrics.LOSS in model.metrics:
                        loss_list = model.values[constants.Metrics.LOSS]
                        reconstruction_error_on_train = loss_list[model.best_epoch - 1]
                    ws.append([
                        str(round(rate * 100, 0)).replace('.', ','),
                        model.stopped_epoch,
                        model.best_epoch,
                        str(reconstruction_error_on_train).replace('.', ','),
                        str(model.monitor_best_value).replace('.', ',')])
        wb.save(path_to_workbook)

    @staticmethod
    def create_autoencoder_classifier_together_report(experiment_results: List[ExperimentResult], path_to_workbook,
                                                      experiment_type):
        if Path(path_to_workbook).exists():
            wb = load_workbook(path_to_workbook)
        else:
            wb = Workbook()
        ws = wb.active
        ws.append(["Typ experimentu", experiment_type])
        ws.append(
            ['Vrstvy enkódera trénovateľné',
             "Percento trénovacích dát autoenkóder",
             "Percento trénovacích dát klasifikátor",
             "Rekonštrukčná chyba na testovacom datasete",
             "Presnosť na trénovacom datasete",
             "Presnosť na testovacom datasete"])
        for result in experiment_results:
            train_history: BasicTrainParametersTwoModels = result.train_history.train_parameters
            rate_autoencoder = train_history.train_parameters_autoencoder.train_data_rate
            rate_classifier = train_history.train_parameters_classifier.train_data_rate
            encoder_trainable = train_history.autoencoder_layers_trainable_during_classifier_training
            encoder_trainable_str = "Áno" if encoder_trainable else "Nie"
            classifier_best_value = None

            accuracy_on_train = -1

            for model in result.train_history.train_history_model_list:
                if model.model_name == constants.Models.CLASSIFIER:
                    classifier_best_value = model.monitor_best_value
                    if constants.Metrics.ACCURACY in model.metrics:
                        accuracy_list = model.values[constants.Metrics.ACCURACY]
                        accuracy_on_train = accuracy_list[model.best_epoch - 1]

            model_with_weights = result.experiment.model_provider()
            instantietied_dataset = result.experiment.dataset_provider()
            import experiments
            autoencoder_test_best_value = experiments.ExperimentAutoencoder.evaluate_on_test(instantietied_dataset,
                                                                                             model_with_weights)
            ws.append([encoder_trainable_str,
                       clear_float(round(rate_autoencoder * 100, 0)),
                       clear_float(round(rate_classifier * 100, 0)),
                       clear_float(autoencoder_test_best_value),
                       clear_float(accuracy_on_train),
                       str(classifier_best_value).replace('.', ',')])
        wb.save(path_to_workbook)

    @staticmethod
    def create_autoclassifier_report(experiment_results: List[ExperimentResult], path_to_workbook,
                                     experiment_type):
        if Path(path_to_workbook).exists():
            wb = load_workbook(path_to_workbook)
        else:
            wb = Workbook()
        ws = wb.active
        ws.append(["Typ experimentu", experiment_type])
        ws.append(
            ["Percento trénovacích dát",
             "Rekonštrukčná chyba na testovacom datasete",
             "Presnosť na trénovacom datasete",
             "Presnosť na testovacom datasete"])
        for result in experiment_results:
            train_history: BasicTrainParametersAutoClassifier = result.train_history.train_parameters
            rate = train_history.train_data_rate
            for model in result.train_history.train_history_model_list:
                if model.model_name == constants.Models.AUTO_CLASSIFIER:
                    accuracy_on_train = -1
                    val_reconstruction_error = -1
                    if constants.Metrics.CLASSIFIER_OUT_ACCURACY in model.metrics:
                        accuracy_list = model.values[constants.Metrics.CLASSIFIER_OUT_ACCURACY]
                        accuracy_on_train = accuracy_list[model.best_epoch - 1]
                    if constants.Metrics.VAL_AUTOENCODER_OUT_LOSS in model.metrics:
                        val_reconstruction_error = model.values[constants.Metrics.VAL_AUTOENCODER_OUT_LOSS][
                            model.best_epoch - 1]
                    ws.append([
                        clear_float(round(rate * 100, 0)),
                        clear_float(val_reconstruction_error),
                        clear_float(accuracy_on_train),
                        clear_float(model.monitor_best_value)])
        wb.save(path_to_workbook)

    @staticmethod
    def create_report(experiment_results: List[ExperimentResult], path_to_workbook,
                      experiment_type):
        if len(experiment_results) > 0:
            result: ExperimentResult = experiment_results[0]
            experiment = result.experiment
            if isinstance(experiment, ExperimentClassifier):
                Reporter.create_classifier_report(experiment_results, path_to_workbook, experiment_type)
            elif isinstance(experiment, ExperimentAutoencoder):
                Reporter.create_autoencoder_report(experiment_results, path_to_workbook, experiment_type)
            elif isinstance(experiment, ExperimentAutoencoderAndClassifier):
                Reporter.create_autoencoder_classifier_together_report(experiment_results, path_to_workbook,
                                                                       experiment_type)
            elif isinstance(experiment, ExperimentAutoClassifier):
                Reporter.create_autoclassifier_report(experiment_results, path_to_workbook,
                                                      experiment_type)
            else:
                raise Exception("Unknown experiment type")

    @staticmethod
    def statistic_for_datasets(mnist_train: np.array,
                               mnist_test: np.array,
                               fashion_mnist_train: np.array,
                               fashion_mnist_test: np.array,
                               cifar_train: np.array,
                               cifar_test: np.array,
                               note: str,
                               path_to_workbook: str):

        from excel_logger import ExcelLogger
        ExcelLogger.init_logger(path_to_workbook)

        def info(label, train, test):
            ExcelLogger.write_into_sheet([label,
                                          clear_float(np.average(train)),
                                          clear_float(np.std(train)),
                                          clear_float(np.average(test)),
                                          clear_float(np.std(test))
                                          ])

        ExcelLogger.write_into_sheet([note])
        ExcelLogger.write_into_sheet(["Dataset",
                                      "Priemer trénovací dataset",
                                      "Smerodajná odchylka trénovací dataset",
                                      "Priemer testovací dataset",
                                      "Smerodajná odchylka testovací dataset"])
        info("Mnist", mnist_train, mnist_test)
        info("Fashion mnist", fashion_mnist_train, fashion_mnist_test)
        # info("Fashion mnist 5 classes", fashion_mnist_five_train, fashion_mnist_five_test)
        info("Cifar", cifar_train, cifar_test)

    @staticmethod
    def statistics_for_autoencoder_unknown(
            original_dataset: np.array,
            original_dataset_name: str,
            mnist_train: np.array,
            mnist_test: np.array,
            fashion_mnist_train: np.array,
            fashion_mnist_test: np.array,
            fashion_mnist_five_train: np.array,
            fashion_mnist_five_test: np.array,
            fashion_mnist_five_other_train: np.array,
            fashion_mnist_five_other_test: np.array,
            cifar_train: np.array,
            cifar_test: np.array,
            path_to_workbook: str):
        from excel_logger import ExcelLogger
        ExcelLogger.init_logger(path_to_workbook)

        def get_unknown(original_list, unknown_list):
            std = np.std(original_list)
            mean = np.mean(original_list)
            unknown = 0
            for num in unknown_list:
                if num < (mean - 2 * std) or num > (mean + 2 * std):
                    unknown += 1
            return unknown

        def info(label, train, test):
            ExcelLogger.write_into_sheet([label,
                                          clear_float(np.average(train)),
                                          clear_float(np.std(train)),
                                          clear_float(np.average(test)),
                                          clear_float(np.std(test))
                                          ])

        if fashion_mnist_five_train is None or fashion_mnist_five_test is None or \
                fashion_mnist_five_other_train is None or fashion_mnist_five_other_test is None:

            ExcelLogger.write_into_sheet(["Neznáme trénovací dataset " + original_dataset_name,
                                          "Neznámy Mnist",
                                          "Neznámy Fashion Mnist",
                                          "Neznámy Cifar Gray",
                                          ])
            ExcelLogger.write_into_sheet([
                clear_float(get_unknown(original_dataset, original_dataset) / len(original_dataset)),
                clear_float(get_unknown(original_dataset, mnist_test) / len(mnist_test)),
                clear_float(get_unknown(original_dataset, fashion_mnist_test) / len(fashion_mnist_test)),
                clear_float(get_unknown(original_dataset, cifar_test) / len(cifar_test))
            ])
        else:
            ExcelLogger.write_into_sheet(["Neznáme trénovací dataset " + original_dataset_name,
                                          "Neznámy Mnist",
                                          "Neznámy Fashion Mnist Five classes",
                                          "Neznámy Fashion Mnist Five other classes",
                                          "Neznámy Cifar Gray",
                                          ])
            ExcelLogger.write_into_sheet([
                clear_float(get_unknown(original_dataset, original_dataset) / len(original_dataset)),
                clear_float(get_unknown(original_dataset, mnist_test) / len(mnist_test)),
                clear_float(get_unknown(original_dataset, fashion_mnist_five_test) / len(fashion_mnist_five_test)),
                clear_float(
                    get_unknown(original_dataset, fashion_mnist_five_other_test) / len(fashion_mnist_five_other_test)),
                clear_float(get_unknown(original_dataset, cifar_test) / len(cifar_test))
            ])

    @staticmethod
    def report_derivated_and_merged_dataset(autoencoder_model: models.Models,
                                            original_model: models.Models,
                                            derivated_model: models.Models,
                                            merged_model: models.Models,
                                            original_dataset: datasets.Dataset,
                                            path_to_excel: str, description: str):
        from excel_logger import ExcelLogger
        ExcelLogger.init_logger(path_to_excel)
        ExcelLogger.write_into_sheet([description])
        ExcelLogger.write_into_sheet(
            ["Classifier", "Original dataset train", "Original dataset test", "Derivated dataset train",
             "Derivated dataset test",
             "Merged dataset train", "Merged dataset test"])
        new_dataset = datasets.DerivatedDataset(original_dataset, autoencoder_model)
        merged_dataset = datasets.MergedDataset(original_dataset, new_dataset)

        for info, model in [('Original', original_model), ('Derivated', derivated_model), ('Merged', merged_model)]:
            original_train = experiments.ExperimentClassifier.evaluate_on_train(original_dataset, model)
            original_test = experiments.ExperimentClassifier.evaluate_on_test(original_dataset, model)

            new_train = experiments.ExperimentClassifier.evaluate_on_train(new_dataset, model)
            new_test = experiments.ExperimentClassifier.evaluate_on_test(new_dataset, model)

            merged_train = experiments.ExperimentClassifier.evaluate_on_train(merged_dataset, model)
            merged_test = experiments.ExperimentClassifier.evaluate_on_test(merged_dataset, model)
            ExcelLogger.write_into_sheet([info,
                                          clear_float(original_train),
                                          clear_float(original_test),
                                          clear_float(new_train),
                                          clear_float(new_test),
                                          clear_float(merged_train),
                                          clear_float(merged_test),
                                          ])

        ExcelLogger.init_logger(None)

    @staticmethod
    def load_last_model(path: str):
        return ExperimentBase.provide_existing_model_from_log(path, ExperimentBase.predicate_for_last_model())()

    @staticmethod
    def report_concrete_derivated_merged_mnist():
        Reporter.report_derivated_and_merged_dataset(
            Reporter.load_last_model(constants.ExperimentsPaths.FashionMnist.BasicModelBuilder.AUTOENCODER),
            Reporter.load_last_model(constants.ExperimentsPaths.FashionMnist.BasicModelBuilder.CLASSIFIER_EN_LAYERS_ON),
            Reporter.load_last_model(
                constants.ExperimentsPaths.FashionMnist.BasicModelBuilder.AUTOENCODER_DERIVATED_CLASSIFIER),
            Reporter.load_last_model(
                constants.ExperimentsPaths.FashionMnist.BasicModelBuilder.AUTOENCODER_MERGED_CLASSIFIER),
            datasets.FashionMnistDataset(),
            'delete.xlsx',
            "Fashion Mnist autoencoder from exp 1 - AUTOENCODER Basic Model Builder"
        )
        Reporter.report_derivated_and_merged_dataset(
            Reporter.load_last_model(
                constants.ExperimentsPaths.FashionMnist.BasicModelBuilder.AUTOENCODER1_CLASSIFIER1_AUTO_L_ON),
            Reporter.load_last_model(constants.ExperimentsPaths.FashionMnist.BasicModelBuilder.CLASSIFIER_EN_LAYERS_ON),
            Reporter.load_last_model(
                constants.ExperimentsPaths.FashionMnist.BasicModelBuilder.AUTOENCODER1_CLASSIFIER1_AUTO_L_ON_DERIVATED_CLASSIFIER),
            Reporter.load_last_model(
                constants.ExperimentsPaths.FashionMnist.BasicModelBuilder.AUTOENCODER1_CLASSIFIER1_AUTO_L_ON_MERGED_CLASSIFIER),
            datasets.FashionMnistDataset(),
            'delete.xlsx',
            "Fashion Mnist autoencoder from exp 1 - AUTOENCODER1_CLASSIFIER1_AUTO_L_ON Basic Model Builder"
        )
        Reporter.report_derivated_and_merged_dataset(
            Reporter.load_last_model(
                constants.ExperimentsPaths.FashionMnist.BasicModelBuilder.AUTO_CLASSIFIER_1to1),
            Reporter.load_last_model(constants.ExperimentsPaths.FashionMnist.BasicModelBuilder.CLASSIFIER_EN_LAYERS_ON),
            Reporter.load_last_model(
                constants.ExperimentsPaths.FashionMnist.BasicModelBuilder.AUTO_CLASSIFIER_1to1_DERIVATED_CLASSIFIER),
            Reporter.load_last_model(
                constants.ExperimentsPaths.FashionMnist.BasicModelBuilder.AUTO_CLASSIFIER_1to1_MERGED_CLASSIFIER),
            datasets.FashionMnistDataset(),
            'delete.xlsx',
            "Fashion Mnist autoencoder from exp 1 - AUTO_CLASSIFIER_1to1 Basic Model Builder"
        )

    @staticmethod
    def report_concrete_derivated_merged_cifargray():
        print("oprava")
        Reporter.report_derivated_and_merged_dataset(
            Reporter.load_last_model(constants.ExperimentsPaths.CifarGray.LargeModelBuilder.AUTOENCODER),
            Reporter.load_last_model(constants.ExperimentsPaths.CifarGray.LargeModelBuilder.CLASSIFIER_EN_LAYERS_ON),
            Reporter.load_last_model(
                constants.ExperimentsPaths.CifarGray.LargeModelBuilder.AUTOENCODER_DERIVATED_CLASSIFIER),
            Reporter.load_last_model(
                constants.ExperimentsPaths.CifarGray.LargeModelBuilder.AUTOENCODER_MERGED_CLASSIFIER),
            datasets.Cifar10GrayDataset(),
            'delete.xlsx',
            "Cifar gray autoencoder from exp 1 - AUTOENCODER Large Model Builder"
        )
        Reporter.report_derivated_and_merged_dataset(
            Reporter.load_last_model(
                constants.ExperimentsPaths.CifarGray.LargeModelBuilder.AUTOENCODER1_CLASSIFIER1_AUTO_L_ON),
            Reporter.load_last_model(constants.ExperimentsPaths.CifarGray.LargeModelBuilder.CLASSIFIER_EN_LAYERS_ON),
            Reporter.load_last_model(
                constants.ExperimentsPaths.CifarGray.LargeModelBuilder.AUTOENCODER1_CLASSIFIER1_AUTO_L_ON_DERIVATED_CLASSIFIER),
            Reporter.load_last_model(
                constants.ExperimentsPaths.CifarGray.LargeModelBuilder.AUTOENCODER1_CLASSIFIER1_AUTO_L_ON_MERGED_CLASSIFIER),
            datasets.Cifar10GrayDataset(),
            'delete.xlsx',
            "Cifar gray autoencoder from exp 1 - AUTOENCODER1_CLASSIFIER1_AUTO_L_ON Large Model Builder"
        )
        Reporter.report_derivated_and_merged_dataset(
            Reporter.load_last_model(constants.ExperimentsPaths.CifarGray.LargeModelBuilder.AUTO_CLASSIFIER_1to1),
            Reporter.load_last_model(constants.ExperimentsPaths.CifarGray.LargeModelBuilder.CLASSIFIER_EN_LAYERS_ON),
            Reporter.load_last_model(
                constants.ExperimentsPaths.CifarGray.LargeModelBuilder.AUTO_CLASSIFIER_1to1_DERIVATED_CLASSIFIER),
            Reporter.load_last_model(
                constants.ExperimentsPaths.CifarGray.LargeModelBuilder.AUTO_CLASSIFIER_1to1_MERGED_CLASSIFIER),
            datasets.Cifar10GrayDataset(),
            'delete.xlsx',
            "Cifar gray autoencoder from exp 1 - AUTO_CLASSIFIER_1to1 Large Model Builder"
        )


def parse_list_from_file(path):
    result = []
    with open(path, 'r') as file:
        for line in file:
            number = float(line.split(";")[1])
            result.append(number)
    return np.array(result)


ROOTROOT = constants.AutoencoderEvaluationPaths.CifarGray.LargeModelBuilder
ROOT = ROOTROOT.Train
mnist_train = parse_list_from_file(ROOT.MNIST_EVAL)
fashion_mnist_train = parse_list_from_file(ROOT.FASHION_MNIST_EVAL)
fashion_mnist_five_train = None  # parse_list_from_file(ROOT.FASHION_MNIST_FIVE_CLASSES_EVAL)
fashion_mnist_five_other_train = None  # parse_list_from_file(ROOT.FASHION_MNIST_FIVE_CLASSES_OTHER_EVAL)
cifar_train = parse_list_from_file(ROOT.CIFAR_GRAY_EVAL)

ROOT = ROOTROOT.Test
mnist_test = parse_list_from_file(ROOT.MNIST_EVAL)
fashion_mnist_test = parse_list_from_file(ROOT.FASHION_MNIST_EVAL)
fashion_mnist_five_test = None  # parse_list_from_file(ROOT.FASHION_MNIST_FIVE_CLASSES_EVAL)
fashion_mnist_five_other_test = None  # parse_list_from_file(ROOT.FASHION_MNIST_FIVE_CLASSES_OTHER_EVAL)
cifar_test = parse_list_from_file(ROOT.CIFAR_GRAY_EVAL)

# Reporter.statistics_for_autoencoder_unknown(
#     cifar_train,
#     "Fashion Mnist Five classes",
#     mnist_train,
#     mnist_test,
#     fashion_mnist_train,
#     fashion_mnist_test,
#     fashion_mnist_five_train,
#     fashion_mnist_five_test,
#     fashion_mnist_five_other_train,
#     fashion_mnist_five_other_test,
#     cifar_train,
#     cifar_test,
#     "delete.xlsx"
# )
#
# Reporter.statistic_for_datasets(
#     mnist_train,
#     mnist_test,
#     fashion_mnist_train,
#     fashion_mnist_test,
#     cifar_train,
#     cifar_test,
#     "Autoenkoder cifar LM",
#     "delete.xlsx"
# )

# graphs
# from matplotlib import pyplot as plt
# from matplotlib.pyplot import figure
# figure(num=None, figsize=(10, 8), dpi=100, facecolor='w', edgecolor='k')
#
# bins = np.linspace(0,0.8 , 20)
#
# plt.hist(mnist_train, bins, alpha=0.5, label='Mnist train')
# plt.hist(fashion_mnist_train, bins, alpha=0.5, label='Fashion mnist train')
# plt.hist(fashion_mnist_five_train, bins, alpha=0.5, label='Fashion mnist 5 classes train')
# plt.hist(cifar_train, bins, alpha=0.5, label='Cifar gray train')
# plt.xlabel("Reconstruction error")
# plt.ylabel("Number of samples")
# plt.legend(loc='upper right')
# plt.show()


if __name__ == "__main__":
    pass
    path = 'example.xlsx'

    # import os
    # if os.path.exists(path):
    #     os.remove(path)
    # results = experiments.ExperimentBase.load_experiment_results(ROOT_PATH.AUTOENCODER, False)
    # results.sort(key= lambda result: result.train_history.train_parameters.train_data_rate)
    # reporter.Reporter.create_report(results, path, "1. Autoenkóder")

    # results = experiments.ExperimentBase.load_experiment_results(ROOT_PATH.CLASSIFIER_EN_LAYERS_ON, False)
    # results.sort(key= lambda result: result.train_history.train_parameters.train_data_rate)
    # reporter.Reporter.create_report(results, path, "2. Klasifikátor - vrstvy enkódera trénovateľné")

    # results = experiments.ExperimentBase.load_experiment_results(ROOT_PATH.CLASSIFIER_EN_LAYERS_OFF, False)
    # results.sort(key= lambda result: result.train_history.train_parameters.train_data_rate)
    # reporter.Reporter.create_report(results, path, "3. Klasifikátor - vrstvy enkódera nie sú trénovateľné")

    # path = 'delete.xlsx'
    # results = experiments.ExperimentBase.load_experiment_results(
    #     constants.ExperimentsPaths.FashionMnist.BasicModelBuilder.AUTOENCODER1_CLASSIFIER1_AUTO_L_ON, True)
    # Reporter.create_report(results, path, "4.Autoenóder a klasifikátor striedavo - vrstvy enkódera sú trénovateľné")

    # Reporter.append_to_workbook(["Cifar gray", "Large Model Builder"], path)
    #
    # results = experiments.ExperimentBase.load_experiment_results(
    #     constants.ExperimentsPaths.CifarGray.LargeModelBuilder.AUTOENCODER, False)
    # Reporter.create_report(results, path, "1. Autoenkóder")
    #
    # results = experiments.ExperimentBase.load_experiment_results(
    #     constants.ExperimentsPaths.CifarGray.LargeModelBuilder.CLASSIFIER_EN_LAYERS_ON, False)
    # Reporter.create_report(results, path, "2. Klasifikátor - vrstvy enkódera trénovateľné")
    #
    # results = experiments.ExperimentBase.load_experiment_results(
    #     constants.ExperimentsPaths.CifarGray.LargeModelBuilder.AUTOENCODER1_CLASSIFIER1_AUTO_L_ON, False)
    # Reporter.create_report(results, path, "4.Autoenóder a klasifikátor striedavo - vrstvy enkódera sú trénovateľné")
    #
    # results = experiments.ExperimentBase.load_experiment_results(
    #     constants.ExperimentsPaths.CifarGray.LargeModelBuilder.AUTOENCODER_TRAINED_CLASSIFIER_AUTO_L_OFF, False)
    # Reporter.create_report(results, path,
    #                        "7. Autoenkóder natrénovaný a dotrénovanie klasifikátora - vrstvy enkódera nie sú trénovateľné")
    #
    # results = experiments.ExperimentBase.load_experiment_results(
    #     constants.ExperimentsPaths.CifarGray.LargeModelBuilder.AUTO_CLASSIFIER_1to1, False)
    # Reporter.create_report(results, path, "8. Autoklasifikátor 1A k 1KL")
    #
    # results = experiments.ExperimentBase.load_experiment_results(
    #     constants.ExperimentsPaths.CifarGray.LargeModelBuilder.AUTO_CLASSIFIER_1to5, False)
    # Reporter.create_report(results, path, "8. Autoklasifikátor 1A k 5KL")
    #
    # results = experiments.ExperimentBase.load_experiment_results(
    #     constants.ExperimentsPaths.CifarGray.LargeModelBuilder.AUTO_CLASSIFIER_5to1, False)
    # Reporter.create_report(results, path, "8. Autoklasifikátor 5A k 1KL")
    #
    # Reporter.append_to_workbook(["Cifar gray five classes", "Large Model Builder"], path)
    #
    # results = experiments.ExperimentBase.load_experiment_results(
    #     constants.ExperimentsPaths.CifarGrayFiveClasses.LargeModelBuilder.CLASSIFIER_EN_LAYERS_ON, False)
    # Reporter.create_report(results, path, "2. Klasifikátor - vrstvy enkódera trénovateľné")
    #
    # results = experiments.ExperimentBase.load_experiment_results(
    #     constants.ExperimentsPaths.CifarGrayFiveClasses.LargeModelBuilder.AUTOENCODER1_CLASSIFIER1_AUTO_L_ON, False)
    # Reporter.create_report(results, path, "4.Autoenóder a klasifikátor striedavo - vrstvy enkódera sú trénovateľné")
    #
    # results = experiments.ExperimentBase.load_experiment_results(
    #     constants.ExperimentsPaths.CifarGrayFiveClasses.LargeModelBuilder.AUTO_CLASSIFIER_1to1, False)
    # Reporter.create_report(results, path, "8. Autoklasifikátor 1A k 1KL")
    # Reporter.report_concrete_derivated_merged_mnist()
